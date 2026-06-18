import asyncio
import json
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import APIRouter, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from src.engines.configs.raw_string_searcher_config import RAW_STRING_RULES
from src.engines.configs.rule import SPELL_CHECK_RULES
from src.engines.raw_searcher import RawStringSearcher
from src.engines.spell_checker import SpellChecker
from src.models.interface import SpellError, Tag
from src.parsers.csv_parser import CsvParser
from src.parsers.excel_parser import ExcelParser
from src.parsers.srt_parser import SrtParser
from src.parsers.txt_parser import TxtParser
from src.tokenizations.ko_tokenizer import KoTokenizer
from src.utils.paths import backend_resource_path, frontend_dist_path, user_data_path

_PARSERS = {
    ".srt": SrtParser,
    ".txt": TxtParser,
    ".xlsx": ExcelParser,
    ".csv": CsvParser,
}

_raw_searcher: RawStringSearcher | None = None
_spell_checker: SpellChecker | None = None
_tokenizer: KoTokenizer | None = None
_tokenizer_lock: asyncio.Lock | None = None

_USER_WORDS_PATH = user_data_path("user_words.json")


# ──────────────────────────────────────────────────────────────
# 사용자 사전 데이터
# ──────────────────────────────────────────────────────────────
def _load_user_data() -> dict:
    if not _USER_WORDS_PATH.exists():
        return {"global": [], "projects": {}, "active_project": None}
    with open(_USER_WORDS_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    # migrate old format (flat list → global words)
    if isinstance(raw, list):
        return {"global": raw, "projects": {}, "active_project": None}
    return raw


def _save_user_data(data: dict) -> None:
    _USER_WORDS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_USER_WORDS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _get_active_words(data: dict) -> list[dict]:
    words = list(data.get("global", []))
    active = data.get("active_project")
    if active:
        words += data.get("projects", {}).get(active, [])
    return words


# ──────────────────────────────────────────────────────────────
# lifespan: 엔진 초기화
# ──────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    global _raw_searcher, _spell_checker, _tokenizer, _tokenizer_lock
    _tokenizer_lock = asyncio.Lock()

    _raw_searcher = RawStringSearcher()
    _raw_searcher.add_word_from_list(RAW_STRING_RULES)

    _tokenizer = KoTokenizer()
    _tokenizer.tokenize("")
    for entry in _get_active_words(_load_user_data()):
        _tokenizer.add_user_word(entry["word"], Tag[entry["tag"]])

    _spell_checker = SpellChecker()
    _spell_checker.add_rule_from_list(SPELL_CHECK_RULES)
    yield


# ──────────────────────────────────────────────────────────────
# 모델
# ──────────────────────────────────────────────────────────────
class CheckRequest(BaseModel):
    text: str


class WordEntry(BaseModel):
    word: str
    tag: str


class AddWordsRequest(BaseModel):
    words: list[WordEntry]
    scope: str = "global"


class ProjectRequest(BaseModel):
    name: str


class ActivateProjectRequest(BaseModel):
    project: str | None = None


class SpellErrorResponse(BaseModel):
    error_type: str
    error_message: str
    start_index: int
    end_index: int
    rule_id: str
    detailed: str


class ExcelConfig(BaseModel):
    sheet_name: str
    text_col: str
    metadata_col: str | None = None
    has_header: bool = True


class CsvConfig(BaseModel):
    text_col: str
    metadata_col: str | None = None
    encoding: str = "utf-8"


class CsvInfoRequest(BaseModel):
    path: str
    encoding: str = "utf-8"


class FileCheckRequest(BaseModel):
    path: str
    excel_config: ExcelConfig | None = None
    csv_config: CsvConfig | None = None


class SegmentResult(BaseModel):
    metadata: str
    text: str
    errors: list[SpellErrorResponse]


class SheetInfo(BaseModel):
    name: str
    columns: list[str]
    has_header: bool = True


# ──────────────────────────────────────────────────────────────
# 공통 헬퍼
# ──────────────────────────────────────────────────────────────
def _to_response(e: SpellError) -> SpellErrorResponse:
    return SpellErrorResponse(
        error_type=e.error_type.name,
        error_message=e.error_message,
        start_index=e.start_index,
        end_index=e.end_index,
        rule_id=e.rule_id,
        detailed=e.detailed,
    )


def _apply_user_words_to_tokenizer(words: list[dict], score: float = 0.0) -> None:
    global _tokenizer
    _tokenizer = KoTokenizer.reset()
    for entry in words:
        _tokenizer.add_user_word(entry["word"], Tag[entry["tag"]], score)


def _col_letter(idx: int) -> str:
    s = ""
    idx += 1
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s

def _detect_has_header(file_path: Path, sheet_name: str) -> bool:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    wb.close()
    if not first_row:
        return False
    return any(c is not None for c in first_row)

# ──────────────────────────────────────────────────────────────
# 라우터 1: check (항상 노출)
# ──────────────────────────────────────────────────────────────
check_router = APIRouter(tags=["check"])


@check_router.post("/raw-check", response_model=list[SpellErrorResponse])
def raw_check(body: CheckRequest) -> list[SpellErrorResponse]:
    assert _raw_searcher is not None
    results: list[SpellError] = _raw_searcher.search(body.text)
    return [_to_response(e) for e in results]


@check_router.post("/nfa-check", response_model=list[SpellErrorResponse])
async def nfa_check(body: CheckRequest) -> list[SpellErrorResponse]:
    assert _spell_checker is not None
    assert _tokenizer is not None
    assert _tokenizer_lock is not None
    async with _tokenizer_lock:
        tokens = _tokenizer.tokenize(body.text)
    results = list(_spell_checker.check(tokens))
    return [_to_response(e) for e in results]


@check_router.post("/check", response_model=list[SpellErrorResponse])
async def check(body: CheckRequest) -> list[SpellErrorResponse]:
    assert _raw_searcher is not None
    assert _spell_checker is not None
    assert _tokenizer is not None
    assert _tokenizer_lock is not None
    raw_results: list[SpellError] = _raw_searcher.search(body.text)
    async with _tokenizer_lock:
        tokens = _tokenizer.tokenize(body.text)
    nfa_results = list(_spell_checker.check(tokens))
    return [_to_response(e) for e in [*raw_results, *nfa_results]]


# ──────────────────────────────────────────────────────────────
# 라우터 2: file (desktop 전용)
# ──────────────────────────────────────────────────────────────
file_router = APIRouter(tags=["file"])


@file_router.post("/excel-info", response_model=list[SheetInfo])
def excel_info(body: FileCheckRequest) -> list[SheetInfo]:
    import openpyxl
    file_path = Path(body.path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result: list[SheetInfo] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        columns = [str(c) for c in first_row if c is not None] if first_row else []

        has_header = bool(columns)
        if not columns:
            ncols = ws.max_column or 0
            columns = [_col_letter(i) for i in range(ncols)]

        result.append(SheetInfo(name=sheet_name, columns=columns, has_header=has_header))
    wb.close()
    return result


@file_router.post("/csv-info", response_model=list[str])
def csv_info(body: CsvInfoRequest) -> list[str]:
    import csv
    file_path = Path(body.path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")
    with open(file_path, encoding=body.encoding, newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
    return headers


@file_router.post("/file-check", response_model=list[SegmentResult])
async def file_check(body: FileCheckRequest) -> list[SegmentResult]:
    assert _raw_searcher is not None
    assert _spell_checker is not None
    assert _tokenizer is not None
    assert _tokenizer_lock is not None

    file_path = Path(body.path)
    ext = file_path.suffix.lower()
    parser_cls = _PARSERS.get(ext)
    if parser_cls is None:
        raise HTTPException(status_code=400, detail=f"지원하지 않는 파일 형식: {ext}")
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")

    if ext == ".xlsx":
        if body.excel_config is None:
            raise HTTPException(status_code=400, detail="엑셀 파일은 excel_config가 필요합니다.")
        cfg = body.excel_config
        parser = ExcelParser(sheet_name=cfg.sheet_name, text_col=cfg.text_col, metadata_col=cfg.metadata_col, has_header=cfg.has_header)
    elif ext == ".csv":
        if body.csv_config is None:
            raise HTTPException(status_code=400, detail="CSV 파일은 csv_config가 필요합니다.")
        cfg = body.csv_config
        parser = CsvParser(text_col=cfg.text_col, metadata_col=cfg.metadata_col, encoding=cfg.encoding)
    else:
        parser = parser_cls()
    segments = list(parser.parse(file_path))
    results: list[SegmentResult] = []

    async with _tokenizer_lock:
        for seg in segments:
            raw_errors: list[SpellError] = _raw_searcher.search(seg.text)
            tokens = _tokenizer.tokenize(seg.text)
            nfa_errors = list(_spell_checker.check(tokens))
            errors = [_to_response(e) for e in [*raw_errors, *nfa_errors]]
            results.append(SegmentResult(metadata=seg.metadata, text=seg.text, errors=errors))

    return results


# ──────────────────────────────────────────────────────────────
# 라우터 3: admin (desktop 전용)
# ──────────────────────────────────────────────────────────────
admin_router = APIRouter(tags=["admin"])


@admin_router.get("/projects")
def get_projects() -> dict:
    data = _load_user_data()
    return {"names": list(data.get("projects", {}).keys()), "active": data.get("active_project")}


@admin_router.post("/projects", status_code=201)
async def create_project(body: ProjectRequest):
    assert _tokenizer_lock is not None
    async with _tokenizer_lock:
        data = _load_user_data()
        if body.name in data.get("projects", {}):
            raise HTTPException(status_code=409, detail="이미 존재하는 프로젝트입니다.")
        data.setdefault("projects", {})[body.name] = []
        _save_user_data(data)
    return {"name": body.name}


@admin_router.delete("/projects/{name}")
async def delete_project(name: str):
    assert _tokenizer_lock is not None
    async with _tokenizer_lock:
        data = _load_user_data()
        if name not in data.get("projects", {}):
            raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
        del data["projects"][name]
        if data.get("active_project") == name:
            data["active_project"] = None
            _apply_user_words_to_tokenizer(_get_active_words(data))
        _save_user_data(data)
    return {"deleted": name}


@admin_router.post("/activate-project")
async def activate_project(body: ActivateProjectRequest):
    assert _tokenizer_lock is not None
    async with _tokenizer_lock:
        data = _load_user_data()
        if body.project is not None and body.project not in data.get("projects", {}):
            raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
        data["active_project"] = body.project
        _save_user_data(data)
        _apply_user_words_to_tokenizer(_get_active_words(data))
    return {"active": body.project}


@admin_router.get("/words")
def get_words(scope: str = "global") -> list[dict]:
    data = _load_user_data()
    if scope == "global":
        return data.get("global", [])
    if scope not in data.get("projects", {}):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    return data["projects"][scope]


@admin_router.post("/add-words", status_code=200)
async def add_words(body: AddWordsRequest):
    assert _tokenizer_lock is not None
    async with _tokenizer_lock:
        data = _load_user_data()
        updated = [e.model_dump() for e in body.words]
        if body.scope == "global":
            data["global"] = updated
        else:
            if body.scope not in data.get("projects", {}):
                raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
            data["projects"][body.scope] = updated
        _save_user_data(data)
        _apply_user_words_to_tokenizer(_get_active_words(data))
    return {"saved": len(updated)}

_FRONTEND_DIST = frontend_dist_path()

def _mount_frontend(app: FastAPI) -> None:
    if not _FRONTEND_DIST.exists():
        return
    app.mount("/assets", StaticFiles(directory=_FRONTEND_DIST / "assets"), name="assets")

    @app.get("/")
    def serve_index():
        return FileResponse(_FRONTEND_DIST / "index.desktop.html")

    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        file = _FRONTEND_DIST / full_path
        if file.exists() and file.is_file():
            return FileResponse(file)
        return FileResponse(_FRONTEND_DIST / "index.desktop.html")


def create_app(mode: str = "web") -> FastAPI:
    """애플리케이션 팩토리.

    mode="desktop": 전체 API(파일/사전 관리 포함) + 프론트 서빙.
    mode="web"    : check 계열만 노출(데모용, 안전 기본값).
    """
    is_desktop = mode.lower() == "desktop"

    app = FastAPI(lifespan=lifespan)

    app.add_middleware(
        CORSMiddleware,
        allow_origins=[
            "http://localhost:5173",
            "https://ink-14.github.io",
        ],
        allow_methods=["GET", "POST"],
        allow_headers=["*"],
    )

    app.include_router(check_router)

    # desktop 전용
    if is_desktop:
        app.include_router(file_router)
        app.include_router(admin_router)
        _mount_frontend(app)

    return app